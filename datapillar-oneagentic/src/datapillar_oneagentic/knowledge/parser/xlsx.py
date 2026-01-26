"""XLSX parser."""

from __future__ import annotations

import io

from datapillar_oneagentic.knowledge.models import DocumentInput, ParsedDocument
from datapillar_oneagentic.knowledge.parser.base import DocumentParser
from datapillar_oneagentic.knowledge.parser.utils import build_document_id, guess_mime_type, load_bytes, normalize_metadata


class XlsxParser(DocumentParser):
    supported_mime_types = {
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    }
    supported_extensions = {".xlsx"}
    name = "xlsx"

    def parse(self, doc_input: DocumentInput) -> ParsedDocument:
        data = load_bytes(doc_input)
        mime_type = guess_mime_type(doc_input)
        content = _xlsx_to_text(data)
        return ParsedDocument(
            document_id=build_document_id(),
            source_type="file",
            mime_type=mime_type,
            text=content,
            pages=[content] if content else [],
            metadata=normalize_metadata(doc_input.metadata),
        )


def _xlsx_to_text(data: bytes) -> str:
    try:
        from openpyxl import load_workbook
    except ImportError as err:
        raise ImportError(
            "XLSX parsing requires dependencies:\n"
            "  pip install datapillar-oneagentic[knowledge]"
        ) from err

    wb = load_workbook(filename=io.BytesIO(data), read_only=True, data_only=True)
    lines = []
    for sheet in wb.worksheets:
        lines.append(f"# Sheet: {sheet.title}")
        for row in sheet.iter_rows(values_only=True):
            row_values = ["" if cell is None else str(cell) for cell in row]
            lines.append("\t".join(row_values).strip())
    return "\n".join([line for line in lines if line])
